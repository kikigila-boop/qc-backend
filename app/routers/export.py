"""
Export router — download QC list as Excel or PDF.
Supports the same filters as GET /qc: search, status, qc_result, editor_name, date_from, date_to.
"""
import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_

from ..database import get_db
from ..models.user import User
from ..models.qc_content import QCContent, StatusEnum
from ..utils.security import get_current_user

router = APIRouter(prefix="/export", tags=["Export"])

# ── shared query builder ─────────────────────────────────────────────────────

def _build_query(
    db: Session,
    search: Optional[str],
    status: Optional[str],
    qc_result: Optional[str],
    editor_name: Optional[str],
    date_from: Optional[datetime],
    date_to: Optional[datetime],
):
    q = db.query(QCContent)
    if search:
        like = f"%{search}%"
        q = q.filter(or_(
            QCContent.qcid.ilike(like),
            QCContent.title.ilike(like),
            QCContent.episode.ilike(like),
            QCContent.season.ilike(like),
            QCContent.editor_name.ilike(like),
        ))
    if status:
        q = q.filter(QCContent.status == status)
    if qc_result:
        q = q.filter(QCContent.qc_result == qc_result)
    if editor_name:
        q = q.filter(QCContent.editor_name.ilike(f"%{editor_name}%"))
    if date_from:
        q = q.filter(QCContent.created_at >= date_from)
    if date_to:
        q = q.filter(QCContent.created_at <= date_to)
    return q.order_by(QCContent.created_at.desc()).all()


# ── column definitions ───────────────────────────────────────────────────────

COLUMNS = [
    ("QCID",         lambda r: r.qcid or f"PENDING-{r.id}"),
    ("Judul",        lambda r: r.title),
    ("Season",       lambda r: r.season or ""),
    ("Episode",      lambda r: r.episode or ""),
    ("Status",       lambda r: r.status.value if hasattr(r.status, "value") else str(r.status)),
    ("QC Result",    lambda r: r.qc_result or ""),
    ("Editor",       lambda r: r.editor_name or ""),
    ("Duration",     lambda r: r.duration or ""),
    ("Storage",      lambda r: r.storage_location or ""),
    ("Ingest By",    lambda r: r.ingest_by or ""),
    ("Ingest At",    lambda r: r.ingest_at.strftime("%Y-%m-%d %H:%M") if r.ingest_at else ""),
    ("QC Date",      lambda r: r.qc_date.strftime("%Y-%m-%d") if r.qc_date else ""),
    ("Catatan",      lambda r: r.notes or ""),
    ("Revised Notes",lambda r: r.revised_notes or ""),
    ("Dibuat",       lambda r: r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else ""),
]


# ── Excel export ─────────────────────────────────────────────────────────────

@router.get("/excel")
def export_excel(
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    qc_result: Optional[str] = Query(None),
    editor_name: Optional[str] = Query(None),
    date_from: Optional[datetime] = Query(None),
    date_to: Optional[datetime] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    rows = _build_query(db, search, status, qc_result, editor_name, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "QC Data"

    # Header style
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [col[0] for col in COLUMNS]
    ws.append(headers)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 32

    # Data rows
    STATUS_COLORS = {
        "QC Process":     "FEF3C7",
        "QC Done":        "D1FAE5",
        "Uploading":      "DBEAFE",
        "Ready To Ingest":"EDE9FE",
        "Done Ingest":    "D1FAE5",
        "Revised":        "FED7AA",
    }
    for row_data in rows:
        values = [getter(row_data) for _, getter in COLUMNS]
        ws.append(values)
        row_num = ws.max_row
        status_val = values[4]
        row_fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status_val, "FFFFFF"))
        for col_idx, cell in enumerate(ws[row_num], 1):
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if col_idx == 5:  # Status column
                cell.fill = row_fill
        ws.row_dimensions[row_num].height = 18

    # Column widths
    widths = [14, 36, 8, 8, 16, 10, 18, 10, 20, 14, 16, 12, 24, 24, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"QC_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── PDF export ───────────────────────────────────────────────────────────────

@router.get("/pdf")
def export_pdf(
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    qc_result: Optional[str] = Query(None),
    editor_name: Optional[str] = Query(None),
    date_from: Optional[datetime] = Query(None),
    date_to: Optional[datetime] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    rows = _build_query(db, search, status, qc_result, editor_name, date_from, date_to)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontSize=14, fontName="Helvetica-Bold", alignment=TA_CENTER)
    sub_style   = ParagraphStyle("sub",   fontSize=8,  fontName="Helvetica",      alignment=TA_CENTER, textColor=colors.grey)
    cell_style  = ParagraphStyle("cell",  fontSize=7,  fontName="Helvetica",      leading=9)

    # Subset of columns for PDF (keep narrow)
    PDF_COLS = [
        ("QCID",      lambda r: r.qcid or f"P-{r.id}"),
        ("Judul",     lambda r: r.title),
        ("S/E",       lambda r: f"{r.season}/{r.episode}"),
        ("Status",    lambda r: r.status.value if hasattr(r.status, "value") else str(r.status)),
        ("Result",    lambda r: r.qc_result or ""),
        ("Editor",    lambda r: r.editor_name or ""),
        ("QC Date",   lambda r: r.qc_date.strftime("%Y-%m-%d") if r.qc_date else ""),
        ("Ingest By", lambda r: r.ingest_by or ""),
        ("Catatan",   lambda r: (r.notes or "")[:60]),
    ]

    header_row = [col[0] for col in PDF_COLS]
    data_rows  = [[Paragraph(getter(r), cell_style) for _, getter in PDF_COLS] for r in rows]
    table_data = [header_row] + data_rows

    col_widths = [2.2*cm, 6.5*cm, 1.5*cm, 2.8*cm, 1.5*cm, 3*cm, 2*cm, 2.5*cm, 4.5*cm]

    STATUS_PDF_COLORS = {
        "QC Process":     colors.HexColor("#FEF3C7"),
        "QC Done":        colors.HexColor("#D1FAE5"),
        "Uploading":      colors.HexColor("#DBEAFE"),
        "Ready To Ingest":colors.HexColor("#EDE9FE"),
        "Done Ingest":    colors.HexColor("#BBFBCA"),
        "Revised":        colors.HexColor("#FED7AA"),
    }

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    ts = TableStyle([
        # Header
        ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#4F46E5")),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0),  7),
        ("ALIGN",       (0, 0), (-1, 0),  "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
        ("FONTSIZE",    (0, 1), (-1, -1), 7),
        ("TOPPADDING",  (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0,0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ])

    # Color status column per row
    for i, row_data in enumerate(rows, 1):
        status_val = row_data.status.value if hasattr(row_data.status, "value") else str(row_data.status)
        bg = STATUS_PDF_COLORS.get(status_val)
        if bg:
            ts.add("BACKGROUND", (3, i), (3, i), bg)

    table.setStyle(ts)

    generated_at = datetime.now().strftime("%d %B %Y, %H:%M")
    filter_desc = " · ".join(filter(None, [
        f"Status: {status}" if status else None,
        f"Result: {qc_result}" if qc_result else None,
        f"Editor: {editor_name}" if editor_name else None,
        f'Cari: "{search}"' if search else None,
    ])) or "Semua data"

    story = [
        Paragraph("OTT QC Management — Export Data", title_style),
        Paragraph(f"Filter: {filter_desc}  |  {len(rows)} item  |  Diekspor: {generated_at}", sub_style),
        Spacer(1, 0.4*cm),
        table,
    ]

    doc.build(story)
    buf.seek(0)

    filename = f"QC_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
