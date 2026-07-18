from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import Optional
from datetime import datetime
import io

from ..database import get_db
from ..models.library import LibraryEntry, LibraryIdCounter

router = APIRouter(prefix="/library", tags=["Library"])

SHOW_TYPES = [
    "Drama", "Comedy", "Action", "Thriller", "Romance", "Horror",
    "Sci-Fi", "Fantasy", "Animation", "Documentary", "Reality",
    "Variety", "Kids", "Sports", "News",
]

GENRE_MAP = {
    "Drama": ["Romance Drama", "Family Drama", "Historical Drama", "Medical Drama", "Legal Drama", "Crime Drama"],
    "Comedy": ["Romantic Comedy", "Sitcom", "Dark Comedy"],
    "Action": ["Martial Arts", "Thriller Action", "Adventure"],
    "Horror": ["Supernatural Horror", "Psychological Horror", "Slasher"],
    "Thriller": ["Thriller - General"],
    "Romance": ["Romance - General"],
    "Sci-Fi": ["Sci-Fi - General"],
    "Fantasy": ["Fantasy - General"],
    "Animation": ["Animation - General"],
    "Documentary": ["Documentary - General"],
    "Reality": ["Reality - General"],
    "Variety": ["Variety - General"],
    "Kids": ["Kids - General"],
    "Sports": ["Sports - General"],
    "News": ["News - General"],
}

RATING_OPTIONS = ["SU", "BO", "R", "D", "A"]
CONTENT_TYPES = ["Microdrama", "Series", "Movies", "Trailer"]
ENTRY_TYPES = ["series", "season", "title", "subtitle", "poster"]

COUNTRY_CODES = {
    "ID": "Indonesia", "CN": "China", "KR": "Korea", "JP": "Japan",
    "TH": "Thailand", "US": "United States", "IN": "India",
    "MY": "Malaysia", "PH": "Philippines",
}

LANGUAGE_CODES = {
    "id": "Indonesian", "en": "English", "zh": "Chinese", "ko": "Korean",
    "ja": "Japanese", "th": "Thai", "hi": "Hindi", "ms": "Malay", "tl": "Filipino",
}

PROVIDERS = [
    "VIU", "VIKI", "WeTV", "iQIYI", "Netflix", "Prime Video",
    "Disney+", "Mango TV", "Youku", "Other",
]

def _entry_to_dict(entry: LibraryEntry) -> dict:
    return {
        "id": entry.id,
        "library_id": entry.library_id,
        "platform": entry.platform,
        "creation_date": entry.creation_date,
        "provider": entry.provider,
        "type": entry.type,
        "show_type": entry.show_type,
        "content_type": entry.content_type,
        "qc_status": entry.qc_status,
        "title_en": entry.title_en,
        "title_id": entry.title_id,
        "summary_long_en": entry.summary_long_en,
        "summary_long_id": entry.summary_long_id,
        "summary_short_en": entry.summary_short_en,
        "summary_short_id": entry.summary_short_id,
        "rating": entry.rating,
        "run_time": entry.run_time,
        "display_run_time": entry.display_run_time,
        "country_of_origin": entry.country_of_origin,
        "genre": entry.genre,
        "actors": entry.actors,
        "directors": entry.directors,
        "producers": entry.producers,
        "studio_name": entry.studio_name,
        "languages": entry.languages,
        "subtitle_languages": entry.subtitle_languages,
        "season_number": entry.season_number,
        "year": entry.year,
        "ingestion_date": entry.ingestion_date,
        "qc_date": entry.qc_date,
        "material_date": entry.material_date,
        "airing_date": entry.airing_date,
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "updated_at": entry.updated_at.isoformat() if entry.updated_at else None,
        "is_complete": entry.is_complete,
    }


def _next_library_id(platform: str, db: Session) -> str:
    label = "VShort" if (platform or "").lower() == "vshort" else "VPlus"
    counter_row = (
        db.query(LibraryIdCounter)
        .filter(LibraryIdCounter.platform == label)
        .with_for_update()
        .first()
    )
    if counter_row is None:
        counter_row = LibraryIdCounter(platform=label, counter=1)
        db.add(counter_row)
    else:
        counter_row.counter += 1
    db.flush()
    return f"{datetime.utcnow().strftime('%Y%m%d')}-{label}-{counter_row.counter:04d}"


@router.get("/reference")
def get_reference_data():
    return {
        "show_types": SHOW_TYPES,
        "genre_map": GENRE_MAP,
        "rating_options": RATING_OPTIONS,
        "content_types": CONTENT_TYPES,
        "entry_types": ENTRY_TYPES,
        "country_codes": COUNTRY_CODES,
        "language_codes": LANGUAGE_CODES,
        "providers": PROVIDERS,
    }


@router.get("/export/excel")
def export_library_excel(
    platform: Optional[str] = Query(None),
    qc_status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    q = db.query(LibraryEntry)
    if platform:
        q = q.filter(LibraryEntry.platform == platform)
    if qc_status:
        q = q.filter(LibraryEntry.qc_status == qc_status)
    entries = q.order_by(LibraryEntry.id.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Library"
    headers = [
        "library_id", "platform", "creation_date", "provider", "type", "show_type",
        "content_type", "qc_status", "title_en", "title_id",
        "summary_long_en", "summary_long_id", "summary_short_en", "summary_short_id",
        "rating", "run_time", "display_run_time", "country_of_origin", "genre",
        "actors", "directors", "producers", "studio_name", "languages",
        "subtitle_languages", "season_number", "year",
        "ingestion_date", "qc_date", "material_date", "airing_date", "is_complete",
    ]
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2F4F7F", end_color="2F4F7F", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = ha
    for ri, entry in enumerate(entries, 2):
        d = _entry_to_dict(entry)
        for ci, h in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=d.get(h))
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 50)
    ws.freeze_panes = "A2"
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    fname = f"library_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/")
def list_library_entries(
    search: Optional[str] = Query(None),
    platform: Optional[str] = Query(None),
    show_type: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    qc_status: Optional[str] = Query(None),
    complete: Optional[bool] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db),
):
    q = db.query(LibraryEntry)
    if search:
        p = f"%{search}%"
        q = q.filter(or_(
            LibraryEntry.title_en.ilike(p),
            LibraryEntry.title_id.ilike(p),
            LibraryEntry.library_id.ilike(p),
            LibraryEntry.provider.ilike(p),
        ))
    if platform:
        q = q.filter(LibraryEntry.platform == platform)
    if show_type:
        q = q.filter(LibraryEntry.show_type == show_type)
    if type:
        q = q.filter(LibraryEntry.type == type)
    if qc_status:
        q = q.filter(LibraryEntry.qc_status == qc_status)
    total = q.count()
    entries = q.order_by(LibraryEntry.id.desc()).offset(skip).limit(limit).all()
    items = [_entry_to_dict(e) for e in entries]
    if complete is not None:
        items = [i for i in items if i["is_complete"] == complete]
        total = len(items)
    return {"items": items, "total": total}


@router.post("/")
def create_library_entry(payload: dict, db: Session = Depends(get_db)):
    platform = payload.get("platform", "vplus")
    library_id = _next_library_id(platform, db)
    entry = LibraryEntry(
        library_id=library_id, platform=platform,
        creation_date=payload.get("creation_date"),
        provider=payload.get("provider"),
        type=payload.get("type"),
        show_type=payload.get("show_type"),
        content_type=payload.get("content_type"),
        qc_status=payload.get("qc_status"),
        title_en=payload.get("title_en"),
        title_id=payload.get("title_id"),
        summary_long_en=payload.get("summary_long_en"),
        summary_long_id=payload.get("summary_long_id"),
        summary_short_en=payload.get("summary_short_en"),
        summary_short_id=payload.get("summary_short_id"),
        rating=payload.get("rating"),
        run_time=payload.get("run_time"),
        display_run_time=payload.get("display_run_time"),
        country_of_origin=payload.get("country_of_origin"),
        genre=payload.get("genre"),
        actors=payload.get("actors"),
        directors=payload.get("directors"),
        producers=payload.get("producers"),
        studio_name=payload.get("studio_name"),
        languages=payload.get("languages"),
        subtitle_languages=payload.get("subtitle_languages"),
        season_number=payload.get("season_number"),
        year=payload.get("year"),
        ingestion_date=payload.get("ingestion_date"),
        qc_date=payload.get("qc_date"),
        material_date=payload.get("material_date"),
        airing_date=payload.get("airing_date"),
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return _entry_to_dict(entry)


@router.patch("/{library_id}")
def update_library_entry(library_id: str, payload: dict, db: Session = Depends(get_db)):
    entry = db.query(LibraryEntry).filter(LibraryEntry.library_id == library_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail=f"Library entry '{library_id}' not found")
    updatable = {
        "platform", "creation_date", "provider", "type", "show_type", "content_type",
        "qc_status", "title_en", "title_id", "summary_long_en", "summary_long_id",
        "summary_short_en", "summary_short_id", "rating", "run_time", "display_run_time",
        "country_of_origin", "genre", "actors", "directors", "producers", "studio_name",
        "languages", "subtitle_languages", "season_number", "year",
        "ingestion_date", "qc_date", "material_date", "airing_date",
    }
    for field, value in payload.items():
        if field in updatable:
            setattr(entry, field, value)
    db.commit()
    db.refresh(entry)
    return _entry_to_dict(entry)
